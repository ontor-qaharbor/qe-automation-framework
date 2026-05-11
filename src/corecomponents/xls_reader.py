from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


class XlsReader:
    """Excel read/write helpers."""

    @staticmethod
    def read_rows(xlsx_path: str | Path, sheet_name: str) -> list[dict[str, Any]]:
        wb = load_workbook(filename=str(xlsx_path), data_only=True)
        ws = wb[sheet_name]
        header = [
            str(c.value).strip() if c.value is not None else ""
            for c in next(ws.iter_rows(min_row=1, max_row=1))
        ]
        rows: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r is None:
                continue
            row = {header[i]: r[i] for i in range(min(len(header), len(r)))}
            rows.append(row)
        return rows


@dataclass(frozen=True)
class LoginRow:
    title: str
    username: str
    password: str


def load_login_rows(xlsx_path: str | Path, sheet_name: str = "login") -> list[LoginRow]:
    wb = load_workbook(filename=str(xlsx_path), data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet {sheet_name!r} not found in {xlsx_path!s}")
    ws = wb[sheet_name]

    header = [
        str(c.value).strip().lower() if c.value is not None else ""
        for c in next(ws.iter_rows(min_row=1, max_row=1))
    ]
    idx = {name: header.index(name) for name in ("title", "username", "password")}

    rows: list[LoginRow] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r is None:
            continue
        title = (
            (r[idx["title"]] or "").strip()
            if isinstance(r[idx["title"]], str)
            else str(r[idx["title"]] or "").strip()
        )
        username = (
            (r[idx["username"]] or "").strip()
            if isinstance(r[idx["username"]], str)
            else str(r[idx["username"]] or "").strip()
        )
        password = (
            (r[idx["password"]] or "").strip()
            if isinstance(r[idx["password"]], str)
            else str(r[idx["password"]] or "").strip()
        )
        if not title and not username and not password:
            continue
        rows.append(LoginRow(title=title, username=username, password=password))

    return rows


def get_login_by_title(xlsx_path: str | Path, title: str, sheet_name: str = "login") -> LoginRow:
    for row in load_login_rows(xlsx_path, sheet_name=sheet_name):
        if row.title == title:
            return row
    raise KeyError(f"Login row not found for title={title!r} in sheet {sheet_name!r}")


def add_login_row(xlsx_path: str | Path, title: str, username: str, password: str, sheet_name: str = "login") -> None:
    wb = load_workbook(filename=str(xlsx_path))
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.append(["title", "username", "password"])
    else:
        ws = wb[sheet_name]
    ws.append([title, username, password])
    wb.save(xlsx_path)

